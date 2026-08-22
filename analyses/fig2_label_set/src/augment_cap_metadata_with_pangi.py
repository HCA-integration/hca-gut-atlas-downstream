#!/usr/bin/env python3
"""Add HGCA taxonomy and curated PanGI correspondence to a CAP workbook."""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_INPUT = None
DEFAULT_OUTPUT = None
DEFAULT_TAXONOMY = REPO_ROOT / "data" / "demo" / "GCA_taxonomy_2026_CAP.csv"
DEFAULT_CROSSWALK = Path()

ANNOTATION_SHEETS = ("all-cells", "epithelial", "lymphoid", "myeloid", "stromal")
LEVEL_COLUMNS = [f"hgca_celltype_level{i}" for i in range(1, 6)]
ADDED_COLUMNS = [
    "hgca_celltype_v0",
    *LEVEL_COLUMNS,
    "pangi_corresponding_labels",
    "has_corresponding_pangi_label",
    "pangi_correspondence_relationships",
    "pangi_correspondence_confidence",
]


def clean(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    return " ".join(str(value).replace("\r", " ").replace("\n", " ").split())


def as_bool(value: object) -> bool:
    return clean(value).lower() in {"true", "t", "1", "yes", "y"}


def joined_unique(values: pd.Series) -> str:
    seen: list[str] = []
    for value in values:
        text = clean(value)
        if text and text not in seen:
            seen.append(text)
    return "; ".join(seen)


def load_sources(
    taxonomy_path: Path, crosswalk_path: Path
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    taxonomy = pd.read_csv(taxonomy_path, dtype=str).fillna("")
    required_tax = ["hgca_celltype_v1", "hgca_celltype_v0", *LEVEL_COLUMNS]
    missing_tax = [col for col in required_tax if col not in taxonomy]
    if missing_tax:
        raise ValueError(f"Taxonomy is missing columns: {missing_tax}")
    for col in required_tax:
        taxonomy[col] = taxonomy[col].map(clean)
    taxonomy_lookup = (
        taxonomy[required_tax]
        .loc[lambda x: x["hgca_celltype_v1"].ne("")]
        .drop_duplicates("hgca_celltype_v1", keep="first")
        .set_index("hgca_celltype_v1")
    )

    crosswalk = pd.read_csv(crosswalk_path, dtype=str).fillna("")
    required_crosswalk = [
        "pangi_level3_label",
        "pangi_n_cells",
        "hgca_v1_label",
        "hgca_v0_label",
        "alternative_hgca_v1_labels",
        "relationship_to_hgca_v1",
        "confidence",
        "include",
        "mapping_notes",
    ]
    missing_crosswalk = [col for col in required_crosswalk if col not in crosswalk]
    if missing_crosswalk:
        raise ValueError(f"Crosswalk is missing columns: {missing_crosswalk}")
    for col in required_crosswalk:
        crosswalk[col] = crosswalk[col].map(clean)
    crosswalk = crosswalk[crosswalk["include"].map(as_bool)].copy()

    valid_v1 = set(taxonomy_lookup.index)
    invalid_v1 = sorted(set(crosswalk["hgca_v1_label"]) - valid_v1)
    if invalid_v1:
        raise ValueError(f"Crosswalk contains HGCA v1 labels absent from taxonomy: {invalid_v1}")

    # Always derive v0 and hierarchy from the selected v1 term so hand-edited
    # primary labels cannot leave stale taxonomy fields.
    for col in ["hgca_celltype_v0", *LEVEL_COLUMNS]:
        target = "hgca_v0_label" if col == "hgca_celltype_v0" else col
        crosswalk[target] = crosswalk["hgca_v1_label"].map(taxonomy_lookup[col])

    crosswalk["has_corresponding_HGCA_v1_label"] = ~crosswalk[
        "relationship_to_hgca_v1"
    ].str.startswith("no direct counterpart")
    crosswalk["mapped_to_HGCA_parent_only"] = crosswalk[
        "relationship_to_hgca_v1"
    ].str.startswith("no direct counterpart")

    primary = (
        crosswalk.sort_values("pangi_n_cells", key=lambda s: pd.to_numeric(s, errors="coerce"), ascending=False)
        .groupby("hgca_v1_label", as_index=True)
        .agg(
            pangi_corresponding_labels=("pangi_level3_label", joined_unique),
            pangi_correspondence_relationships=("relationship_to_hgca_v1", joined_unique),
            pangi_correspondence_confidence=("confidence", joined_unique),
        )
    )
    return taxonomy_lookup, crosswalk, primary


def last_populated_header_column(ws) -> int:
    populated = [cell.column for cell in ws[1] if clean(cell.value)]
    if not populated:
        raise ValueError(f"Sheet {ws.title!r} has no header row")
    return max(populated)


def clone_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def augment_annotation_sheet(ws, taxonomy_lookup: pd.DataFrame, primary: pd.DataFrame) -> None:
    headers = {clean(cell.value): cell.column for cell in ws[1] if clean(cell.value)}
    if "hgca_celltype_v1" not in headers:
        raise ValueError(f"Sheet {ws.title!r} lacks hgca_celltype_v1")
    label_col = headers["hgca_celltype_v1"]
    start_col = last_populated_header_column(ws) + 1
    style_col = start_col - 1

    for offset, header in enumerate(ADDED_COLUMNS):
        col = start_col + offset
        cell = ws.cell(row=1, column=col, value=header)
        clone_cell_style(ws.cell(row=1, column=style_col), cell)
        cell.font = copy(cell.font)
        cell.font = Font(
            name=cell.font.name or "Arial",
            size=cell.font.sz,
            bold=True,
            italic=cell.font.italic,
            color=cell.font.color,
        )
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(col)].width = 24

    missing_labels: list[str] = []
    for row in range(2, ws.max_row + 1):
        label = clean(ws.cell(row=row, column=label_col).value)
        if not label:
            continue
        if label not in taxonomy_lookup.index:
            missing_labels.append(label)
            continue
        tax = taxonomy_lookup.loc[label]
        pan = primary.loc[label] if label in primary.index else None
        values = [
            clean(tax["hgca_celltype_v0"]),
            *[clean(tax[col]) for col in LEVEL_COLUMNS],
            "" if pan is None else clean(pan["pangi_corresponding_labels"]),
            pan is not None,
            "" if pan is None else clean(pan["pangi_correspondence_relationships"]),
            "" if pan is None else clean(pan["pangi_correspondence_confidence"]),
        ]
        for offset, value in enumerate(values):
            target = ws.cell(row=row, column=start_col + offset, value=value)
            clone_cell_style(ws.cell(row=row, column=style_col), target)
            target.alignment = Alignment(wrap_text=True, vertical="top")

    if missing_labels:
        raise ValueError(
            f"Sheet {ws.title!r} has labels missing from newest taxonomy: "
            f"{sorted(set(missing_labels))}"
        )

    if ws.auto_filter.ref:
        ws.auto_filter.ref = (
            f"A1:{get_column_letter(start_col + len(ADDED_COLUMNS) - 1)}{ws.max_row}"
        )


def add_crosswalk_sheet(wb, crosswalk: pd.DataFrame) -> None:
    title = "PanGI-HGCA crosswalk"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)

    columns = [
        "pangi_level3_label",
        "pangi_n_cells",
        "has_corresponding_HGCA_v1_label",
        "mapped_to_HGCA_parent_only",
        "hgca_v1_label",
        "hgca_v0_label",
        *LEVEL_COLUMNS,
        "alternative_hgca_v1_labels",
        "relationship_to_hgca_v1",
        "confidence",
        "mapping_notes",
        "review_status",
    ]
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(col_idx)].width = (
            18 if header == "pangi_n_cells" else 28
        )

    for row_idx, (_, record) in enumerate(crosswalk.iterrows(), start=2):
        for col_idx, header in enumerate(columns, start=1):
            value = record.get(header, "")
            if header in {
                "has_corresponding_HGCA_v1_label",
                "mapped_to_HGCA_parent_only",
            }:
                value = bool(value)
            elif header == "pangi_n_cells":
                value = int(float(value)) if clean(value) else None
            else:
                value = clean(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"


def document_new_fields(wb) -> None:
    title = "titles and descriptions"
    if title not in wb.sheetnames:
        return
    ws = wb[title]
    existing = {
        clean(ws.cell(row=row, column=1).value)
        for row in range(1, ws.max_row + 1)
    }
    descriptions = {
        "hgca_celltype_v0": "Direct HGCA v0 term aligned to hgca_celltype_v1 in the newest CAP taxonomy; blank where v1 has no direct v0 term.",
        "hgca_celltype_level1-5": "Hierarchical HGCA taxonomy levels 1 through 5 from GCA_taxonomy_2026_CAP.csv.",
        "pangi_corresponding_labels": "Curated PanGI level-3 labels whose primary correspondence is this HGCA v1 term; multiple labels are separated by semicolons.",
        "has_corresponding_pangi_label": "TRUE when at least one included PanGI level-3 label maps primarily to this HGCA v1 term.",
        "has_corresponding_HGCA_v1_label": "On the PanGI-HGCA crosswalk sheet, TRUE when PanGI has a direct/partial/composite HGCA v1 correspondence; FALSE when only a broad HGCA parent is available.",
    }
    for field, description in descriptions.items():
        if field in existing:
            continue
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=field)
        ws.cell(row=row, column=2, value=description)
        for col in range(1, ws.max_column + 1):
            clone_cell_style(ws.cell(row=max(2, row - 1), column=col), ws.cell(row=row, column=col))
            ws.cell(row=row, column=col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--taxonomy", type=Path, default=DEFAULT_TAXONOMY)
    parser.add_argument("--crosswalk", type=Path, default=DEFAULT_CROSSWALK)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    for path in (args.input, args.taxonomy, args.crosswalk):
        if not path.is_file():
            raise FileNotFoundError(path)
    if not args.output.parent.is_dir():
        raise FileNotFoundError(args.output.parent)
    if args.output.resolve() == args.input.resolve():
        raise ValueError("Output must be a new workbook, not the source workbook")

    taxonomy_lookup, crosswalk, primary = load_sources(
        args.taxonomy.resolve(), args.crosswalk.resolve()
    )
    wb = load_workbook(args.input.resolve())
    for sheet in ANNOTATION_SHEETS:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Workbook is missing required sheet: {sheet}")
        augment_annotation_sheet(wb[sheet], taxonomy_lookup, primary)
    add_crosswalk_sheet(wb, crosswalk)
    document_new_fields(wb)
    wb.save(args.output.resolve())

    print(f"Saved: {args.output.resolve()}")
    print(f"Annotation sheets augmented: {len(ANNOTATION_SHEETS)}")
    print(f"PanGI crosswalk rows: {len(crosswalk)}")
    print(
        "PanGI labels with direct/partial/composite HGCA correspondence: "
        f"{int(crosswalk['has_corresponding_HGCA_v1_label'].sum())}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
