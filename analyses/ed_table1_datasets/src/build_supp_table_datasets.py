#!/usr/bin/env python3
"""Build HGCA v1 dataset-level supplementary table.

Starts from metadata/tier_1_dataset.csv, restricts to dataset_ids present in
hgca_all_lineages_v1.h5ad, and enriches with live atlas counts plus curated DOIs.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import anndata as ad
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ROOT = Path(__file__).resolve().parents[1]
REPO = Path(__file__).resolve().parents[3]
H5AD = REPO / "data" / "demo" / "hgca_all_lineages_v1_demo.h5ad"
TIER1 = Path()
DOIS = Path()

DATA = ROOT / "data"
OUT = ROOT / "out"
LOGS = ROOT / "logs"

# Map current object dataset_id -> tier_1_dataset.csv row(s)
TIER1_ALIASES: dict[str, list[str]] = {
    "BasuHelmsley": ["BasuHelmsley", "BasuGCARNA"],
    "HamiltonHelmsley": ["HamiltonHelmsley", "KarakashevaHelmlsey"],
    "Wells2025": ["Wells2025", "DominguezUnpub"],
    "Wells2025_2": ["Wells2025_2", "DominguezUnpub2"],
    "Elmentaite2020": ["Elmentaite2020", "Elmentaite2020_3p", "Elmentaite2020_5p"],
}

CONSORTIUM_CONTRIBUTED = {
    "ArendsHelmsley",
    "BasuHelmsley",
    "HamiltonHelmsley",
}

PI_DISPLAY = {
    "Arends,Mark": "Arends, M. J.",
    "anindita,basu": "Basu, A.",
    "Magness,Scott": "Magness, S.",
    "sarah,teichmann": "Teichmann, S. A.",
    "sarah,a,teichmann": "Teichmann, S. A.",
    "liza,konnikova": "Konnikova, L.",
    "karakasheva,tatiana": "Hamilton, K. E. / Karakasheva, T.",
    "guo,zhiyong": "Guo, Z.",
    "yuxia,zhang": "Zhang, Y.",
    "marco,colonna": "Colonna, M.",
    "alison,simmons": "Simmons, A.",
    "ramnik,xavier": "Xavier, R. J.",
    "anderson,carl": "Anderson, C.",
    "Park,Woong-Yang; Tejpar,Sabine": "Park, W.-Y. / Tejpar, S.",
    "Park,Woong-Yang": "Park, W.-Y.",
    "Tejpar,Sabine": "Tejpar, S.",
    "kai,w,wucherpfennig": "Wucherpfennig, K. W.",
    "peng,qiu": "Qiu, P.",
    "ephraim,kenigsberg": "Kenigsberg, E.",
    "saurabh,mehandru": "Mehandru, S.",
    "yeguang,chen": "Chen, Y.",
    "gray,camp": "Camp, J. G.",
    "kean,leslie": "Kean, L. S.",
}

TISSUE_FIX = {
    "ascending_colon": "ascending colon",
    "descending_colon": "descending colon",
    "sigmoid_colon": "sigmoid colon",
    "transverse_colon": "transverse colon",
    "gastrointestinal_system_mesentery": "gastrointestinal system mesentery",
}


def _as_str_series(s: pd.Series) -> pd.Series:
    if hasattr(s, "cat"):
        s = s.astype(str)
    return s.astype(str)


def uniq_join(values, *, n: int | None = None, sep: str = "; ") -> str:
    vals = []
    seen = set()
    for x in values:
        if pd.isna(x):
            continue
        t = str(x).strip()
        if t in ("", "nan", "None", "NA", "not_applicable"):
            continue
        t = TISSUE_FIX.get(t, t)
        if t not in seen:
            seen.add(t)
            vals.append(t)
    vals = sorted(vals, key=str.lower)
    if n is not None and len(vals) > n:
        return sep.join(vals[:n]) + f"{sep}(+{len(vals) - n} more)"
    return sep.join(vals)


def clean_pi(raw: str) -> str:
    if not raw or raw in ("nan", "None"):
        return ""
    if raw in PI_DISPLAY:
        return PI_DISPLAY[raw]
    # multi-value
    if "; " in raw:
        return "; ".join(clean_pi(p) for p in raw.split("; "))
    if "," in raw and raw.lower() == raw:
        parts = [p.strip() for p in raw.split(",") if p.strip()]
        if len(parts) >= 2:
            last = parts[-1].title()
            initials = " ".join(f"{p[0].upper()}." for p in parts[:-1] if p)
            return f"{last}, {initials}"
    if "," in raw:
        last, first = [x.strip() for x in raw.split(",", 1)]
        if first and " " not in first and len(first) > 1:
            return f"{last}, {first[0].upper()}."
        return f"{last}, {first}"
    return raw


def short_citation(row: pd.Series) -> str:
    title = str(row.get("publication_title") or "").strip()
    year = row.get("publication_year")
    journal = str(row.get("publication_journal") or "").strip()
    status = str(row.get("publication_status") or "")
    did = str(row["dataset_id"])
    if status == "unpublished" or not title or title.lower().startswith("arends lab") or title.lower().startswith("basu lab"):
        return f"{did} (unpublished / consortium contributed)"
    # FirstAuthorYear from dataset id stem
    stem = re.sub(r"_(3p|5p|v2|v3|2)$", "", did)
    cite = stem
    if pd.notna(year) and str(year) not in ("", "nan"):
        # avoid duplicating year if already in stem
        if not re.search(r"20\d{2}", stem):
            cite = f"{stem} ({int(year)})"
    if journal:
        cite = f"{cite}; {journal}"
    return cite


def load_tier1_for_dataset(tier1: pd.DataFrame, dataset_id: str) -> pd.Series:
    if tier1.empty or "dataset_id" not in tier1.columns:
        return pd.Series(dtype=object)
    aliases = TIER1_ALIASES.get(dataset_id, [dataset_id])
    hits = tier1[tier1["dataset_id"].astype(str).isin(aliases)]
    if hits.empty:
        return pd.Series(dtype=object)
    # prefer exact id match
    exact = hits[hits["dataset_id"].astype(str) == dataset_id]
    row = exact.iloc[0] if not exact.empty else hits.iloc[0]
    return row


def summarize_object(h5ad_path: Path) -> pd.DataFrame:
    a = ad.read_h5ad(h5ad_path, backed="r")
    cols = [
        c
        for c in [
            "dataset_id",
            "donor_id",
            "sample_id",
            "tissue_level_1",
            "tissue_ontology_term",
            "assay_ontology_term",
            "suspension_type",
            "sample_collection_method",
            "sample_source",
            "age_range",
            "sampled_site_condition",
            "disease_ontology_term",
            "radial_tissue_term",
            "institute",
            "study_pi",
            "publication_doi",
        ]
        if c in a.obs.columns
    ]
    df = a.obs[cols].copy()
    for c in df.columns:
        df[c] = _as_str_series(df[c])

    rows = []
    for did, g in df.groupby("dataset_id", observed=True):
        rows.append(
            {
                "dataset_id": did,
                "n_cells": int(len(g)),
                "n_donors": int(g["donor_id"].nunique()),
                "n_samples": int(g["sample_id"].nunique()),
                "anatomical_regions": uniq_join(g["tissue_level_1"]) if "tissue_level_1" in g else "",
                "tissue_ontology_terms": uniq_join(g["tissue_ontology_term"]) if "tissue_ontology_term" in g else "",
                "assay": uniq_join(g["assay_ontology_term"]) if "assay_ontology_term" in g else "",
                "suspension_type": uniq_join(g["suspension_type"]) if "suspension_type" in g else "",
                "sample_collection_method": uniq_join(g["sample_collection_method"]) if "sample_collection_method" in g else "",
                "sample_source": uniq_join(g["sample_source"]) if "sample_source" in g else "",
                "age_ranges": uniq_join(g["age_range"]) if "age_range" in g else "",
                "site_condition": uniq_join(g["sampled_site_condition"]) if "sampled_site_condition" in g else "",
                "disease_terms": uniq_join(g["disease_ontology_term"]) if "disease_ontology_term" in g else "",
                "radial_layers": uniq_join(g["radial_tissue_term"]) if "radial_tissue_term" in g else "",
                "institute_obs": uniq_join(g["institute"], n=8) if "institute" in g else "",
                "study_pi_obs": uniq_join(g["study_pi"], n=4) if "study_pi" in g else "",
                "doi_obs": uniq_join(g["publication_doi"], n=3) if "publication_doi" in g else "",
            }
        )
    a.file.close()
    return pd.DataFrame(rows)


def build_table(h5ad_path: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    summ = summarize_object(h5ad_path)
    tier1 = pd.read_csv(TIER1) if TIER1.is_file() else pd.DataFrame()
    dois = pd.read_csv(DOIS) if DOIS.is_file() else pd.DataFrame()

    if dois.empty or "dataset_id" not in dois.columns:
        doi_map = pd.DataFrame()
    else:
        doi_map = dois.drop_duplicates("dataset_id").set_index("dataset_id")

    records = []
    for _, s in summ.iterrows():
        did = s["dataset_id"]
        t1 = load_tier1_for_dataset(tier1, did)
        drow = doi_map.loc[did] if (not doi_map.empty and did in doi_map.index) else None

        doi = ""
        title = ""
        year = np.nan
        journal = ""
        status = ""
        if drow is not None:
            doi = "" if pd.isna(drow.get("doi")) or str(drow.get("doi")) == "NA" else str(drow["doi"])
            title = "" if pd.isna(drow.get("title")) else str(drow["title"])
            year = drow.get("year")
            journal = "" if pd.isna(drow.get("journal")) else str(drow["journal"])
            status = "" if pd.isna(drow.get("status")) else str(drow["status"])
        if not doi and s["doi_obs"] and s["doi_obs"] not in ("not_applicable",):
            doi = s["doi_obs"].split("; ")[0]

        pi_raw = s["study_pi_obs"] or (str(t1.get("study_pi")) if len(t1) else "")
        pi = clean_pi(pi_raw)

        if did in CONSORTIUM_CONTRIBUTED or status == "unpublished":
            if doi and status != "unpublished":
                provenance = "Consortium contributed (Helmsley / HCA Gut); now published"
            else:
                provenance = "Consortium contributed (Helmsley / HCA Gut)"
        else:
            provenance = "Published study"

        rec = {
            "dataset_id": did,
            "short_citation": None,  # fill below
            "publication_title": title,
            "publication_year": year if pd.notna(year) else "",
            "publication_journal": journal,
            "publication_doi": doi,
            "publication_status": status if status else ("unpublished" if did in CONSORTIUM_CONTRIBUTED else "resolved"),
            "study_pi": pi,
            "institute": s["institute_obs"],
            "assay": s["assay"],
            "suspension_type": s["suspension_type"],
            "sample_collection_method": s["sample_collection_method"],
            "sample_source": s["sample_source"],
            "anatomical_regions": s["anatomical_regions"],
            "radial_layers": s["radial_layers"],
            "age_ranges": s["age_ranges"],
            "site_condition": s["site_condition"],
            "disease_terms_in_atlas": s["disease_terms"],
            "n_donors": s["n_donors"],
            "n_samples": s["n_samples"],
            "n_cells": s["n_cells"],
            "reference_genome": str(t1.get("reference_genome") or "") if len(t1) else "",
            "alignment_software": str(t1.get("alignment_software") or "") if len(t1) else "",
            "gene_annotation_version": str(t1.get("gene_annotation_version") or "") if len(t1) else "",
            "tier1_description": str(t1.get("description") or "") if len(t1) else "",
            "tier1_comments": str(t1.get("comments") or "") if len(t1) else "",
            "provenance": provenance,
            "tissue_ontology_terms": s["tissue_ontology_terms"],
        }
        # scrub nan strings from tier1
        for k in ("reference_genome", "alignment_software", "gene_annotation_version", "tier1_description", "tier1_comments"):
            if rec[k] in ("nan", "None"):
                rec[k] = ""
        records.append(rec)

    full = pd.DataFrame(records)
    full["short_citation"] = full.apply(short_citation, axis=1)
    full = full.sort_values("dataset_id").reset_index(drop=True)

    # Publication-facing column order / pretty names
    pretty = full.rename(
        columns={
            "dataset_id": "Dataset ID",
            "short_citation": "Study",
            "publication_title": "Publication title",
            "publication_year": "Year",
            "publication_journal": "Journal",
            "publication_doi": "DOI",
            "study_pi": "Study PI / lab",
            "institute": "Institute",
            "assay": "Assay",
            "suspension_type": "Suspension type",
            "sample_collection_method": "Sample collection method",
            "sample_source": "Sample source",
            "anatomical_regions": "Anatomical regions (in HGCA v1)",
            "radial_layers": "Radial layers (in HGCA v1)",
            "age_ranges": "Donor age ranges (years)",
            "site_condition": "Sampled-site condition",
            "disease_terms_in_atlas": "Disease ontology terms (cells retained)",
            "n_donors": "Donors (n)",
            "n_samples": "Samples (n)",
            "n_cells": "Cells (n)",
            "reference_genome": "Reference genome",
            "alignment_software": "Alignment software",
            "gene_annotation_version": "Gene annotation version",
            "provenance": "Provenance",
            "tier1_comments": "Notes",
        }
    )

    pub_cols = [
        "Dataset ID",
        "Study",
        "Publication title",
        "Year",
        "Journal",
        "DOI",
        "Study PI / lab",
        "Institute",
        "Assay",
        "Suspension type",
        "Sample collection method",
        "Sample source",
        "Anatomical regions (in HGCA v1)",
        "Radial layers (in HGCA v1)",
        "Donor age ranges (years)",
        "Sampled-site condition",
        "Disease ontology terms (cells retained)",
        "Donors (n)",
        "Samples (n)",
        "Cells (n)",
        "Reference genome",
        "Alignment software",
        "Gene annotation version",
        "Provenance",
        "Notes",
    ]
    pretty = pretty[pub_cols]

    dictionary = pd.DataFrame(
        [
            ("Dataset ID", "Identifier used in HGCA v1 integrated object (`dataset_id`)."),
            ("Study", "Short study label for manuscript cross-reference."),
            ("Publication title", "Primary publication title when available."),
            ("Year", "Publication year."),
            ("Journal", "Journal or preprint server."),
            ("DOI", "Digital object identifier; blank for unpublished consortium contributions."),
            ("Study PI / lab", "Principal investigator / contributing lab (cleaned from metadata)."),
            ("Institute", "Contributing institute values present on cells in HGCA v1."),
            ("Assay", "10x assay chemistry present in the retained cells."),
            ("Suspension type", "Cell vs nucleus suspension."),
            ("Sample collection method", "Biopsy and/or surgical resection."),
            ("Sample source", "Donor source class (e.g. surgical, living organ, postmortem)."),
            ("Anatomical regions (in HGCA v1)", "Collapsed `tissue_level_1` values among retained cells."),
            ("Radial layers (in HGCA v1)", "Radial tissue compartments represented."),
            ("Donor age ranges (years)", "Age bins among retained donors."),
            ("Sampled-site condition", "Healthy vs adjacent-to-disease labels on retained samples."),
            ("Disease ontology terms (cells retained)", "Disease ontology terms still present after healthy/adjacent filtering."),
            ("Donors (n)", "Unique `donor_id` count in HGCA v1."),
            ("Samples (n)", "Unique `sample_id` count in HGCA v1."),
            ("Cells (n)", "Cell count in HGCA v1 after integration QC."),
            ("Reference genome", "From HCA tier-1 dataset metadata sheet."),
            ("Alignment software", "From HCA tier-1 dataset metadata sheet."),
            ("Gene annotation version", "From HCA tier-1 dataset metadata sheet."),
            ("Provenance", "Published study vs consortium-contributed Helmsley/HCA data."),
            ("Notes", "Tier-1 comments (e.g. remapped by PanGI)."),
        ],
        columns=["Column", "Description"],
    )
    return pretty, dictionary, full


def write_xlsx(pretty: pd.DataFrame, dictionary: pd.DataFrame, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Datasets"

    header_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=8)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    for r_idx, row in enumerate(dataframe_to_rows(pretty, index=False, header=True), 1):
        ws.append(row)
        for c_idx, _ in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx)
            cell.border = thin
            cell.alignment = wrap
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = body_font
                if r_idx % 2 == 0:
                    cell.fill = alt_fill

    # Reasonable widths for publication browsing
    widths = {
        "A": 18,  # Dataset ID
        "B": 28,  # Study
        "C": 42,  # Title
        "D": 8,
        "E": 18,
        "F": 28,
        "G": 22,
        "H": 28,
        "I": 12,
        "J": 12,
        "K": 18,
        "L": 22,
        "M": 28,
        "N": 16,
        "O": 18,
        "P": 16,
        "Q": 28,
        "R": 10,
        "S": 10,
        "T": 10,
        "U": 14,
        "V": 22,
        "W": 14,
        "X": 28,
        "Y": 24,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30
    for i in range(2, ws.max_row + 1):
        ws.row_dimensions[i].height = 36

    # Totals row note
    note = wb.create_sheet("README", 0)
    note["A1"] = "HGCA v1 — Supplementary dataset table"
    note["A1"].font = Font(name="Arial", size=12, bold=True)
    lines = [
        "",
        "One row per dataset_id retained in hgca_all_lineages_v1.h5ad (27 datasets; 24 studies).",
        "Counts (donors / samples / cells) are from the final integrated atlas object, not the original publications.",
        "Source tables: metadata/tier_1_dataset.csv + byTheNumbers/plots/hgca_v1_included_studies_dois.csv + live obs from the h5ad.",
        "Rebuild: python src/build_supp_table_datasets.py",
        "",
        f"Totals: {pretty['Donors (n)'].sum()} donor-rows summed across datasets (not unique across studies); "
        f"{pretty['Samples (n)'].sum()} samples; {pretty['Cells (n)'].sum():,} cells.",
    ]
    for i, line in enumerate(lines, 2):
        note[f"A{i}"] = line
        note[f"A{i}"].font = Font(name="Arial", size=9)
    note.column_dimensions["A"].width = 120

    ws2 = wb.create_sheet("Column dictionary")
    for r_idx, row in enumerate(dataframe_to_rows(dictionary, index=False, header=True), 1):
        ws2.append(row)
        for c_idx in range(1, 3):
            cell = ws2.cell(r_idx, c_idx)
            cell.font = header_font if r_idx == 1 else body_font
            if r_idx == 1:
                cell.fill = header_fill
            cell.alignment = wrap
            cell.border = thin
    ws2.column_dimensions["A"].width = 36
    ws2.column_dimensions["B"].width = 90

    wb.save(path)


def main() -> None:
    global DATA, OUT, LOGS
    parser = argparse.ArgumentParser()
    parser.add_argument("--h5ad", type=Path, default=H5AD)
    parser.add_argument("--outdir", type=Path, default=OUT)
    args = parser.parse_args()
    if "demo" in args.h5ad.name:
        print("DEMO MODE: results are for software checking, not manuscript figures.")
    OUT = args.outdir
    DATA = args.outdir
    LOGS = args.outdir

    DATA.mkdir(parents=True, exist_ok=True)
    OUT.mkdir(parents=True, exist_ok=True)
    LOGS.mkdir(parents=True, exist_ok=True)

    pretty, dictionary, full = build_table(args.h5ad)

    full_path = DATA / "datasets_internal_full.csv"
    pretty_csv = OUT / "SuppTable_HGCA_v1_datasets.csv"
    pretty_xlsx = OUT / "SuppTable_HGCA_v1_datasets.xlsx"
    dict_csv = OUT / "SuppTable_HGCA_v1_datasets_column_dictionary.csv"

    full.to_csv(full_path, index=False)
    pretty.to_csv(pretty_csv, index=False)
    dictionary.to_csv(dict_csv, index=False)
    write_xlsx(pretty, dictionary, pretty_xlsx)

    summary = (
        f"Wrote {len(pretty)} datasets\n"
        f"  cells={pretty['Cells (n)'].sum():,}  "
        f"sample-rows={pretty['Samples (n)'].sum()}  "
        f"donor-rows={pretty['Donors (n)'].sum()}\n"
        f"  {pretty_csv}\n  {pretty_xlsx}\n"
    )
    (LOGS / "build.log").write_text(summary)
    print(summary)


if __name__ == "__main__":
    main()
